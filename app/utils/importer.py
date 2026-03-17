"""Excel importer for NANUK catalog data.

Two import modes:
  1. import_cases_catalog() — preferred, merges:
       - nanuk_cases_only.csv  (filter: which SKUs to include)
       - Case_catalog.xlsx      (mm dimensions, volume)
       - Price Agreement xlsx   (pricing data)
  2. import_excel()            — legacy, Price Agreement only
"""
import csv
import re
import openpyxl
from pathlib import Path

# ── Price Agreement column map ────────────────────────────────────────────────
PRICE_MAP = {
    'upc': 'upc',
    'upc codes': 'upc',
    'new item number': 'sku',
    'old item number': 'old_item_number',
    'item description': 'description',
    'us map price': 'us_map_price',
    'price': 'price',
    'interior dim. (in)': 'dim_interior',
    'interior dim. (in) l x w x h': 'dim_interior',
    'exterior dim. (in)': 'dim_exterior',
    'exterior dim. (in) l x w x h': 'dim_exterior',
}

# ── Case_catalog column map ───────────────────────────────────────────────────
CASE_MAP = {
    'new item number': 'sku',
    'int. length (mm)': 'int_length_mm',
    'int. width (mm)': 'int_width_mm',
    'int. height (mm)': 'int_height_mm',
    'int. dim. (mm) l x w x h': 'dim_interior_mm',
    'ext. length (mm)': 'ext_length_mm',
    'ext. width (mm)': 'ext_width_mm',
    'ext. height (mm)': 'ext_height_mm',
    'ext. dim. (mm) l x w x h': 'dim_exterior_mm',
    'int. dim. (in) l x w x h': 'dim_interior',
    'ext. dim. (in) l x w x h': 'dim_exterior',
}


def _parse_price(val):
    if val is None:
        return None
    s = str(val).strip().replace('$', '').replace(',', '')
    if s.lower() in ('n/a', '', 'none', '-', 'nan'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_mm(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _find_header_row(ws, key_col: str):
    """Find the row where key_col appears and return (data_start_1based, col_map)."""
    key_lower = key_col.lower()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(str(c).strip().lower() == key_lower for c in row if c is not None):
            # Build col map from this row
            col_map = {}
            for ci, val in enumerate(row):
                if val is None:
                    continue
                k = str(val).strip().lower()
                if k in PRICE_MAP and 'sku' in [PRICE_MAP.get(k2) for k2 in PRICE_MAP]:
                    col_map[PRICE_MAP[k]] = ci
                # also try case map
            return row_idx, row
    return None, None


def _build_col_map(header_row, mapping: dict) -> dict:
    """Build {db_field: col_index} from a header row tuple."""
    col_map = {}
    for ci, val in enumerate(header_row):
        if val is None:
            continue
        k = str(val).strip().lower()
        if k in mapping:
            db_field = mapping[k]
            if db_field not in col_map:
                col_map[db_field] = ci
    return col_map


def _find_header(ws, mapping: dict, key_field: str = 'sku'):
    """Scan sheet for a header row containing key_field, return (data_start, col_map)."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_map = _build_col_map(row, mapping)
        if key_field in col_map:
            # Scan next 2 rows for extra sub-headers
            for offset in (1, 2):
                try:
                    sub = list(ws.iter_rows(
                        min_row=row_idx + offset,
                        max_row=row_idx + offset,
                        values_only=True
                    ))[0]
                except IndexError:
                    break
                sub_map = _build_col_map(sub, mapping)
                for fld, ci in sub_map.items():
                    if fld not in col_map:
                        col_map[fld] = ci

            # Find actual data start (skip sub-header rows)
            sku_idx = col_map[key_field]
            data_start = row_idx + 1
            for probe in range(1, 5):
                try:
                    r = list(ws.iter_rows(
                        min_row=row_idx + probe,
                        max_row=row_idx + probe,
                        values_only=True
                    ))[0]
                except IndexError:
                    break
                v = r[sku_idx] if sku_idx < len(r) else None
                if v and str(v).strip() and str(v).strip().lower() not in ('none', 'nan'):
                    data_start = row_idx + probe
                    break
            return data_start, col_map

    return None, {}


# ── Case SKU detection ────────────────────────────────────────────────────────
# Actual case shells start with: T20, T30, T35, 225, 310, 320, 330, 903..996
# Accessories start with: 1-, 30-, 40-, 50-, 55-, 60-, N1x, S9xx, HANDLING
_CASE_SKU_RE = re.compile(r'^(T\d{2,3}[A-Z]|\d{3}[A-Z0-9])', re.IGNORECASE)

def _is_case_sku(sku: str) -> bool:
    """Return True if the SKU pattern indicates a case shell (not an accessory)."""
    return bool(_CASE_SKU_RE.match(sku.strip()))


def _parse_dim_in(val) -> tuple:
    """Parse inch dim string '14.25x6.1x11.1' → (L_mm, W_mm, H_mm) or (None,None,None)."""
    if not val:
        return None, None, None
    s = str(val).strip()
    if s.lower() in ('n/a', 'none', '', 'nan'):
        return None, None, None
    parts = re.split(r'\s*[xX×]\s*', s)
    if len(parts) >= 3:
        try:
            l = float(parts[0]) * 25.4
            w = float(parts[1]) * 25.4
            h = float(parts[2]) * 25.4
            return l, w, h
        except (ValueError, TypeError):
            pass
    return None, None, None


# ── Main: import cases catalog ────────────────────────────────────────────────

def import_cases_catalog(
    cases_csv_path: str,
    catalog_xlsx_path: str,
    price_xlsx_path: str,
    session,
) -> tuple[int, int]:
    """
    Import ALL case shells from Price Agreement + dimensions from Case_catalog.
    Non-case accessories (foam inserts, straps, etc.) are stored with category='other'.

    Strategy:
      1. Price Agreement → primary data source (SKU, description, price, inch dims)
      2. Case_catalog.xlsx → precise mm dimensions for cases that appear there
      3. Cases without catalog dims → convert inch dims from Price Agreement to mm
      4. SKU pattern (_is_case_sku) determines category: 'case' vs 'other'

    Returns (case_count, other_count).
    """
    from app.models import Item, Stock

    # 1. Read Case_catalog.xlsx → {sku: dim_data}  (precise mm)
    dim_data: dict[str, dict] = {}
    wb_cat = openpyxl.load_workbook(catalog_xlsx_path, data_only=True)
    ws_cat = wb_cat.worksheets[0]
    data_start, col_map_cat = _find_header(ws_cat, CASE_MAP, 'sku')

    if data_start is not None and 'sku' in col_map_cat:
        sku_idx = col_map_cat['sku']
        for row in ws_cat.iter_rows(min_row=data_start, values_only=True):
            raw = row[sku_idx] if sku_idx < len(row) else None
            if not raw:
                continue
            sku = str(raw).strip()
            if not sku or sku.lower() in ('none', 'nan'):
                continue

            def _get(field, _row=row):
                idx = col_map_cat.get(field)
                return _row[idx] if idx is not None and idx < len(_row) else None

            int_l = _parse_mm(_get('int_length_mm'))
            int_w = _parse_mm(_get('int_width_mm'))
            int_h = _parse_mm(_get('int_height_mm'))
            ext_l = _parse_mm(_get('ext_length_mm'))
            ext_w = _parse_mm(_get('ext_width_mm'))
            ext_h = _parse_mm(_get('ext_height_mm'))
            vol = round(ext_l * ext_w * ext_h / 1_000_000_000, 6) if ext_l and ext_w and ext_h else None

            dim_data[sku] = {
                'int_length_mm': int_l,
                'int_width_mm':  int_w,
                'int_height_mm': int_h,
                'ext_length_mm': ext_l,
                'ext_width_mm':  ext_w,
                'ext_height_mm': ext_h,
                'volume_m3':     vol,
                'dim_interior':  _get('dim_interior'),
                'dim_exterior':  _get('dim_exterior'),
            }

    # 2. Read Price Agreement xlsx — ALL rows
    wb_pr = openpyxl.load_workbook(price_xlsx_path, data_only=True)
    ws_pr = wb_pr.worksheets[0]
    data_start_pr, col_map_pr = _find_header(ws_pr, PRICE_MAP, 'sku')

    if data_start_pr is None or 'sku' not in col_map_pr:
        raise ValueError("Could not find 'New Item Number' column in Price Agreement.")

    sku_idx_pr = col_map_pr['sku']
    cases_imported = 0
    others_imported = 0
    seen: set[str] = set()

    for row in ws_pr.iter_rows(min_row=data_start_pr, values_only=True):
        raw = row[sku_idx_pr] if sku_idx_pr < len(row) else None
        if not raw:
            continue
        sku = str(raw).strip()
        if not sku or sku.lower() in ('none', 'nan', '') or len(sku) > 40:
            continue
        if sku in seen:
            continue
        seen.add(sku)

        def _gp(field, _row=row):
            idx = col_map_pr.get(field)
            return _row[idx] if idx is not None and idx < len(_row) else None

        desc_raw = _gp('description')
        description = str(desc_raw).strip() if desc_raw else None

        us_map = _parse_price(_gp('us_map_price'))
        price  = _parse_price(_gp('price'))
        if price is None:
            price = us_map

        category = 'case' if _is_case_sku(sku) else 'other'

        # Dimensions: prefer catalog mm; fallback = parse Price Agreement inch dims
        dd = dim_data.get(sku, {})
        int_l = dd.get('int_length_mm')
        int_w = dd.get('int_width_mm')
        int_h = dd.get('int_height_mm')
        ext_l = dd.get('ext_length_mm')
        ext_w = dd.get('ext_width_mm')
        ext_h = dd.get('ext_height_mm')
        vol   = dd.get('volume_m3')

        if not ext_l:
            # Try parsing inch dims from Price Agreement
            int_l_in, int_w_in, int_h_in = _parse_dim_in(_gp('dim_interior'))
            ext_l_in, ext_w_in, ext_h_in = _parse_dim_in(_gp('dim_exterior'))
            if int_l_in:
                int_l, int_w, int_h = int_l_in, int_w_in, int_h_in
            if ext_l_in:
                ext_l, ext_w, ext_h = ext_l_in, ext_w_in, ext_h_in
                vol = round(ext_l * ext_w * ext_h / 1_000_000_000, 6)

        # dim_interior/exterior string (inches from Price Agreement or catalog)
        dim_int_str = dd.get('dim_interior') or _gp('dim_interior')
        dim_ext_str = dd.get('dim_exterior') or _gp('dim_exterior')

        def _clean_dim(v):
            s = str(v).strip() if v else ''
            return s if s.lower() not in ('n/a', 'none', '', 'nan') else None

        item = session.get(Item, sku)
        if item is None:
            item = Item(sku=sku)
            session.add(item)
        if session.get(Stock, sku) is None:
            session.add(Stock(sku=sku, qty_on_hand=0))

        item.upc             = str(_gp('upc')).strip() if _gp('upc') else None
        item.old_item_number = str(_gp('old_item_number')).strip() if _gp('old_item_number') else None
        item.description     = description
        item.us_map_price    = us_map
        item.price           = price
        item.category        = category
        item.dim_interior    = _clean_dim(dim_int_str)
        item.dim_exterior    = _clean_dim(dim_ext_str)
        item.int_length_mm   = int_l
        item.int_width_mm    = int_w
        item.int_height_mm   = int_h
        item.ext_length_mm   = ext_l
        item.ext_width_mm    = ext_w
        item.ext_height_mm   = ext_h
        item.volume_m3       = vol

        if category == 'case':
            cases_imported += 1
        else:
            others_imported += 1

    session.commit()
    return cases_imported, others_imported


# ── Legacy: Price Agreement only ──────────────────────────────────────────────

def import_excel(filepath: str, session) -> tuple[int, int]:
    """Import items from Price Agreement Excel only. Returns (imported, skipped)."""
    from app.models import Item, Stock

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    data_start, col_map = _find_header(ws, PRICE_MAP, 'sku')
    if data_start is None or 'sku' not in col_map:
        raise ValueError("Could not find 'New Item Number' column. Check Excel format.")

    sku_idx = col_map['sku']
    seen_skus: set[str] = set()
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        sku_raw = row[sku_idx] if sku_idx < len(row) else None
        if not sku_raw:
            skipped += 1
            continue
        sku = str(sku_raw).strip()
        if not sku or sku.lower() in ('none', 'nan', '') or len(sku) > 40:
            skipped += 1
            continue
        if sku in seen_skus:
            skipped += 1
            continue
        seen_skus.add(sku)

        def get_val(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        item = session.get(Item, sku)
        if item is None:
            item = Item(sku=sku)
            session.add(item)
            if session.get(Stock, sku) is None:
                session.add(Stock(sku=sku, qty_on_hand=0))

        item.upc             = str(get_val('upc')).strip() if get_val('upc') else None
        item.old_item_number = str(get_val('old_item_number')).strip() if get_val('old_item_number') else None
        item.description     = str(get_val('description')).strip() if get_val('description') else None
        item.us_map_price    = _parse_price(get_val('us_map_price'))
        item.price           = _parse_price(get_val('price'))

        dim_int = get_val('dim_interior')
        item.dim_interior = (
            str(dim_int).strip()
            if dim_int and str(dim_int).strip().lower() not in ('n/a', 'none', '')
            else None
        )
        dim_ext = get_val('dim_exterior')
        item.dim_exterior = (
            str(dim_ext).strip()
            if dim_ext and str(dim_ext).strip().lower() not in ('n/a', 'none', '')
            else None
        )
        imported += 1

    session.commit()
    return imported, skipped
