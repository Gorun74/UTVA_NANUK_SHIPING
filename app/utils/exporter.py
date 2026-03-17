"""Excel export for container orders."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from app.models import Container, ContainerLine, Item


def export_container_order(container_id: int, session) -> io.BytesIO:
    container = session.get(Container, container_id)
    if not container:
        raise ValueError(f"Container {container_id} not found")

    lines = session.execute(
        select(ContainerLine, Item)
        .join(Item, ContainerLine.sku == Item.sku)
        .where(ContainerLine.container_id == container_id)
        .order_by(ContainerLine.id)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Container {container_id}"

    # Styles
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1A3A5C")
    meta_font = Font(size=10)
    total_font = Font(bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(row, col, value, font=None, fill=None, align="left", number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if number_format:
            c.number_format = number_format
        return c

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"NANUK — Container Order #{container_id}"
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header info block
    info = [
        ("Date Ordered:", container.date_ordered),
        ("ETA:", container.expected_arrival_date),
        ("Status:", container.status.upper()),
        ("Container:", f"{container.container_size} ({container.container_fill})"),
        ("USD/AUD Rate:", f"{container.used_rate:.4f}"),
    ]
    if container.ship_total_aud:
        info.append(("Est. Total Shipping:", f"AUD {container.ship_total_aud:,.2f}"))
    if container.total_cbm:
        info.append(("Total CBM:", f"{container.total_cbm:.2f} m³"))

    row = 2
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = Font(size=10)
        row += 1

    # Blank row
    row += 1

    # Column headers
    headers = ["SKU", "Description", "Ext. Dimensions", "Volume (m³)",
               "Qty Ordered", "Unit Price (USD)", "Line Total (USD)"]
    for col, h in enumerate(headers, 1):
        _cell(row, col, h, font=header_font, fill=header_fill, align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # Data rows
    total_qty = 0
    total_usd = 0.0
    alt_fill = PatternFill("solid", fgColor="EEF3F8")

    for idx, (cl, item) in enumerate(lines):
        fill = alt_fill if idx % 2 == 1 else None
        qty = cl.qty_ordered or 0
        unit_usd = cl.unit_price_usd or 0
        line_total = qty * unit_usd

        _cell(row, 1, cl.sku, fill=fill)
        _cell(row, 2, item.description if item else cl.sku, fill=fill)
        _cell(row, 3, item.dim_exterior if item else "", fill=fill)
        _cell(row, 4, item.volume_m3 if item else None, fill=fill, align="center", number_format="0.0000")
        _cell(row, 5, qty, fill=fill, align="center")
        _cell(row, 6, unit_usd, fill=fill, align="right", number_format='"$"#,##0.00')
        _cell(row, 7, line_total, fill=fill, align="right", number_format='"$"#,##0.00')

        total_qty += qty
        total_usd += line_total
        row += 1

    # Totals row
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    _cell(row, 1, "TOTAL", font=total_font, fill=total_fill)
    _cell(row, 2, "", fill=total_fill)
    _cell(row, 3, "", fill=total_fill)
    _cell(row, 4, "", fill=total_fill)
    _cell(row, 5, total_qty, font=total_font, fill=total_fill, align="center")
    _cell(row, 6, "", fill=total_fill)
    _cell(row, 7, total_usd, font=total_font, fill=total_fill, align="right", number_format='"$"#,##0.00')

    # AUD equivalent
    row += 1
    total_aud = total_usd * (container.used_rate or 1)
    ws.cell(row=row, column=6, value="Total AUD equiv.:").font = Font(bold=True, size=10)
    c = ws.cell(row=row, column=7, value=total_aud)
    c.font = Font(bold=True, size=10)
    c.number_format = '"A$"#,##0.00'

    # Column widths
    widths = [20, 40, 25, 12, 12, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
